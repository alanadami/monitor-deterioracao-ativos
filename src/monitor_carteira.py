import yfinance as yf
import pandas as pd
import joblib
from datetime import datetime
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# ======================================================
# CONFIGURAÇÃO
# ======================================================

ATIVOS = [
    "TRIS3.SA", "SAPR4.SA", "KLBN3.SA", "UNIP6.SA", "PETR4.SA",
    "CPFE3.SA", "ITSA4.SA", "SANB11.SA", "BBAS3.SA", "ABCB4.SA",
    "WEGE3.SA", "GGBR4.SA", "VALE3.SA"
]

IBOV = "^BVSP"
DATA_INICIO = "2000-01-01"
JANELA = 24

FEATURES = [
    "retorno",
    "ret_3m",
    "ret_6m",
    "vol_6m",
    "zscore_lag1",
    "zscore_rel_lag1"
]

PASTA_OUTPUT = "outputs"
PASTA_MODELOS = "modelos"

os.makedirs(PASTA_OUTPUT, exist_ok=True)

HOJE = datetime.today().strftime("%Y_%m_%d")

# ======================================================
# COLETA DE PREÇOS
# ======================================================

def baixar_precos(tickers):
    df = yf.download(
        tickers,
        start=DATA_INICIO,
        auto_adjust=True,
        progress=False
    )["Close"]

    df = df[df > 0]
    df.index = pd.to_datetime(df.index)

    return df.resample("ME").last()

# ======================================================
# FEATURES + LABEL
# ======================================================

def calcular_features_e_label(precos_ativos, precos_ibov):
    ret = precos_ativos.pct_change()
    ret_ibov = precos_ibov.pct_change()

    # Z-score absoluto
    media = ret.rolling(JANELA).mean()
    std = ret.rolling(JANELA).std()
    z = (ret - media) / std

    # Z-score relativo
    ret_rel = ret.sub(ret_ibov, axis=0)
    media_rel = ret_rel.rolling(JANELA).mean()
    std_rel = ret_rel.rolling(JANELA).std()
    z_rel = (ret_rel - media_rel) / std_rel

    # Label
    sinal_fraco = (z < 0).astype(int)
    label = ((sinal_fraco == 1) & (sinal_fraco.shift(1) == 1)).astype(int)

    df = pd.DataFrame({
        "retorno": ret.iloc[-1],
        "ret_3m": (1 + ret).rolling(3).apply(lambda x: x.prod() - 1).iloc[-1],
        "ret_6m": (1 + ret).rolling(6).apply(lambda x: x.prod() - 1).iloc[-1],
        "vol_6m": ret.rolling(6).std().iloc[-1],
        "zscore_lag1": z.shift(1).iloc[-1],
        "zscore_rel_lag1": z_rel.shift(1).iloc[-1],
        "label": label.iloc[-1]
    })

    df.index.name = "Ativo"
    return df.reset_index(), precos_ativos.index[-1]

# ======================================================
# MODELO
# ======================================================

def aplicar_modelo(df):
    scaler = joblib.load(f"{PASTA_MODELOS}/scaler.pkl")
    model = joblib.load(f"{PASTA_MODELOS}/modelo_logistico.pkl")

    X = scaler.transform(df[FEATURES])
    df["score_ml"] = model.predict_proba(X)[:, 1]

    return df

# ======================================================
# DIAGNÓSTICO
# ======================================================

def diagnostico(row):
    if row["label"] == 1:
        return "Regime fraco"
    elif row["score_ml"] >= 0.6:
        return "Atenção"
    else:
        return "Normal"

# ======================================================
# ABA INÍCIO
# ======================================================

def criar_aba_inicio(writer):
    texto = [
        "Monitor de Deterioração de Ativos – Carteira de Investimentos",
        "",
        "Sistema de monitoramento defensivo de ações brasileiras.",
        "O objetivo é identificar regimes persistentes de baixo desempenho,",
        "apoiando decisões de manutenção, redução ou pausa de aportes.",
        "",
        "Características:",
        "- Dados mensais de preços ajustados (Yahoo Finance)",
        "- Retornos mensais simples",
        "- Z-score absoluto e relativo ao Ibovespa (24 meses)",
        "- Regime definido por persistência estatística",
        "- Regressão Logística usada como escore auxiliar",
        "",
        "Uso recomendado:",
        "- Avaliação mensal",
        "- Não utilizar para trading",
        "- Não gera recomendações automáticas",
        "",
        "Autor: Alan Alves",
        "Contato: galves.alan@gmail.com"
    ]

    pd.DataFrame({"": texto}).to_excel(writer, sheet_name="INÍCIO", index=False)

# ======================================================
# HISTÓRICO ACUMULADO
# ======================================================

def atualizar_historico(resumo, data_ref):
    caminho = f"{PASTA_OUTPUT}/historico_monitoramento.xlsx"

    hist_mes = resumo[["Ativo", "Score", "Diagnóstico", "label"]].copy()
    hist_mes["Data"] = data_ref
    hist_mes = hist_mes[["Data", "Ativo", "Score", "Diagnóstico", "label"]]

    if os.path.exists(caminho):
        hist = pd.read_excel(caminho)
        hist = hist[
            ~((hist["Data"] == data_ref) &
              (hist["Ativo"].isin(hist_mes["Ativo"])))
        ]
        hist = pd.concat([hist, hist_mes], ignore_index=True)
    else:
        hist = hist_mes

    hist.to_excel(caminho, index=False)

# ======================================================
# EXECUÇÃO PRINCIPAL
# ======================================================

def main():
    print("Baixando preços...")
    precos_ativos = baixar_precos(ATIVOS)
    precos_ibov = baixar_precos([IBOV])[IBOV]

    print("Calculando features...")
    df, data_ref = calcular_features_e_label(precos_ativos, precos_ibov)

    print("Aplicando modelo...")
    df = aplicar_modelo(df)
    df["diagnostico"] = df.apply(diagnostico, axis=1)

    # =========================
    # ABA DADOS
    # =========================
    dados = df.copy()

    # =========================
    # ABA RESUMO
    # =========================
    resumo = df[[
        "Ativo", "diagnostico", "score_ml",
        "retorno", "ret_6m", "label"
    ]].rename(columns={
        "diagnostico": "Diagnóstico",
        "score_ml": "Score",
        "retorno": "Retorno mês",
        "ret_6m": "Retorno 6m"
    })

    ordem = {"Regime fraco": 0, "Atenção": 1, "Normal": 2}
    resumo["ordem"] = resumo["Diagnóstico"].map(ordem)
    resumo = resumo.sort_values(["ordem", "Score"], ascending=[True, False])
    resumo = resumo.drop(columns="ordem")

    arquivo = f"{PASTA_OUTPUT}/monitor_carteira_{HOJE}.xlsx"

    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        criar_aba_inicio(writer)
        resumo.drop(columns="label").to_excel(writer, sheet_name="RESUMO", index=False)
        dados.to_excel(writer, sheet_name="DADOS", index=False)

    atualizar_historico(resumo, data_ref)

    # =========================
# FORMATAÇÃO EXCEL (POR FAIXA)
# =========================

    wb = load_workbook(arquivo)
    ws = wb["RESUMO"]

    # Cores
    verde = PatternFill("solid", fgColor="C6EFCE")
    amarelo = PatternFill("solid", fgColor="FFEB9C")
    vermelho = PatternFill("solid", fgColor="FFC7CE")

    for row in range(2, ws.max_row + 1):

        # -------------------------
        # Diagnóstico (coluna B)
        # -------------------------
        diag_cell = ws[f"B{row}"]
        if diag_cell.value == "Normal":
            diag_cell.fill = verde
        elif diag_cell.value == "Atenção":
            diag_cell.fill = amarelo
        else:  # Regime fraco
            diag_cell.fill = vermelho

        # -------------------------
        # Score (coluna C)
        # -------------------------
        score_cell = ws[f"C{row}"]
        score = score_cell.value

        if score is not None:
            if score < 0.10:
                score_cell.fill = verde
            elif score < 0.30:
                score_cell.fill = amarelo
            else:
                score_cell.fill = vermelho

            score_cell.number_format = "0.000"

    # -------------------------
    # Percentuais
    # -------------------------
    for col in ["D", "E"]:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = "0.00%"

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    wb.save(arquivo)


# ======================================================
# ENTRY POINT
# ======================================================

if __name__ == "__main__":
    main()
